import pandas as pd
import openpyxl as opxl
import tkinter as tk
import datetime
import json
import global_func as gf
from year_chart import Year_Chart
from excel_manager import Excel_Manager
from log import Log

# Optional: GUI Opens asking which file to open.
# Load the log in tracker.
# - Each item contains: $$, Date, Category, Description
# Load year sheet if not empty.

# GUI:
# Previous year button      |grid|      Next year button
#                           |grid|
# View log                  |grid|      Add transaction
#                           |grid|
#                           |grid|

# Year buttons change which sheet is being shown. Tiny and close to grid.
# View log opens the list of transactions
# Add payment:
# - Asks if it is a input or output.
# - Amount (only pos)
# - Date (default is today)
# - Category (rent, groceries, car, credit card, subscriptions, dining, recreational, gifts, misc)
# - Description

path = "./Tracker.xlsx"

# def createYear(logSheet, yearSheet)

excelManager = Excel_Manager(path)

log = Log(excelManager)

gf.check_years(excelManager)

currentYear = Year_Chart(excelManager=excelManager, log=log, year=gf.today.year)
if (str(currentYear.year) not in excelManager.workbook.sheetnames):
    currentYear.create_chart()

# if (curYear in workbook.sheetnames):
#     yearSheet = workbook[curYear]
# else:
#     yearSheet = workbook.create_sheet(str(curYear))

window = tk.Tk()
window.title("Excel Viewer")

window.columnconfigure(0,weight=1)
window.columnconfigure(1,weight=3)
window.columnconfigure(2,weight=1)

window.rowconfigure(0, weight=1)
window.rowconfigure(1, weight=10)
window.rowconfigure(2, weight=1)

gf.setupScreenSize(window)

leftFrame = gf.create_widget(window, tk.Frame)
leftFrame.grid(row=1, column=0, pady=10)
leftFrame.rowconfigure(0, weight=1)
leftFrame.rowconfigure(1, weight=1)

rightFrame = gf.create_widget(window, tk.Frame)
rightFrame.grid(row=1, column=2, pady=10)
rightFrame.rowconfigure(0, weight=1)
rightFrame.rowconfigure(1, weight=1)

middleFrame = gf.create_widget(window, tk.Frame)
middleFrame.grid(row=1, column=1, pady=10)
middleFrame.rowconfigure(0, weight=1)
middleFrame.rowconfigure(1, weight=10)
middleFrame.rowconfigure(2, weight=1)

addTButton = gf.create_widget(rightFrame, tk.Button, text="Add Transaction", command= lambda: gf.prompt_for_transaction(window, path, log, False))
addTButton.grid(row=0, pady=10)

editEButton = gf.create_widget(rightFrame, tk.Button, text="Edit Expectations", command=lambda: gf.prompt_for_expectations(window))
editEButton.grid(row=1, pady=10)

viewLogButton = gf.create_widget(leftFrame, tk.Button, text="View Log", command=lambda: log.display_log(window))
viewLogButton.grid(row=0, pady=10)

editCatButton = gf.create_widget(leftFrame,tk.Button, text="Edit Categories")
editCatButton.grid(row=1, pady=10)

chartLabel = gf.create_widget(middleFrame, tk.Label, textvariable=tk.StringVar(value=str(currentYear.year)), height=1)
chartLabel.grid(row=0, column=1)

previousChartButton = gf.create_widget(middleFrame, tk.Button, text="<", command=lambda: currentYear.previous_year(chartLabel))
previousChartButton.grid(row=0, column=0)

nextChartButton = gf.create_widget(middleFrame, tk.Button, text=">", command=lambda: currentYear.next_year(chartLabel))
nextChartButton.grid(row=0, column=2)


mainChart = currentYear.display_chart(middleFrame, row=1, column=1, columnSize=100, stretch=0)


window.mainloop()